from os import path, walk
import logging
import unicodedata
from contextlib import suppress
from openpyxl import load_workbook
from settings import IMPORT_PROGRESS_FILE, statuses_mapping


def get_candidates(basedir, db_file):
	_, file_extension = path.splitext(db_file)
	if not path.exists(basedir):
		logging.error('Path not exist')
		exit(1)
	if not path.exists(path.join(basedir, db_file)):
		logging.error('Excel file not exist')
		exit(1)
	if file_extension != '.xlsx':
		logging.error('Not .xlsx file')
		exit(1)

	cv_filenames = get_files_int_dir(basedir)
	ready_candidates = set()
	if path.exists(path.join(basedir, IMPORT_PROGRESS_FILE)):
		ready_candidates = get_ready_candidates(basedir)
	all_candidates = get_candidates_from_xls(path.join(basedir, db_file))
	if cv_filenames is None or all_candidates is None:
		return None

	# берём только тех, которые ещё не добавлены
	nesesarry_candidates = list()
	for candidate in all_candidates:
		if candidate['name'] not in ready_candidates:
			nesesarry_candidates.append(candidate)

	for candidate in nesesarry_candidates:
		for cv in cv_filenames:
			if candidate['name'] in cv:
				candidate['cv'] = cv_filenames[cv]
				break
		if 'cv' not in candidate:
			logging.error(f'Не найден файл резюме для кандидата '
						  f'{candidate["name"]} на дожность {candidate["position"]}.')
			return None
	return nesesarry_candidates


def get_files_int_dir(basedir: str) -> dict:
	"""рекурсивно сканируем папки для поиска всех файлов
	возвращаем словарь со всеми файлами в формате {имя файла: путь}"""
	result = dict()
	for currentpath, folders, files in walk(basedir):
		for file in files:
			filepath = path.realpath(path.join(currentpath, file))
			file = unicodedata.normalize('NFC', file)
			result[file] = filepath
	return result


def get_ready_candidates(basedir) -> set:
	"""возвращаем set с именами кандидатов добавленных перед падением скрипта"""
	candidates = set()
	with open(path.join(basedir, IMPORT_PROGRESS_FILE), 'r') as f:
		for line in f:
			candidates.add(line.replace('\n', ''))
	return candidates


def get_candidates_from_xls(filepath):
	xls_db = load_workbook(filename=filepath)
	sheet = xls_db.active
	candidates = []
	for row in sheet.iter_rows(min_row=2, values_only=True):
		position, name, salary, comment, status = row
		name = ' '.join(name.split())  # чистим мусор
		status_ = statuses_mapping.get(status, None)
		if status is None:
			logging.error(f'Status "{status}" not in mapping table')
			return None
		candidates.append(
			{
				'position': position,
				'name': name,
				'salary': salary,
				'comment': comment,
				'status': status_
			}
		)
	return candidates


def add_vacancy_id(canditates: list, vacancies: list):
	for canditate in canditates:
		for vacancy in vacancies:
			if canditate['position'] == vacancy['position']:
				canditate['vacancy_id'] = vacancy['id']
		if 'vacancy_id' not in canditate:
			logging.error(f"В вакансиях нет вакансии на позицию {canditate['position']}")
			return None
	return canditates


def add_status_id(canditates: list, statuses: list):
	for canditate in canditates:
		for status in statuses:
			if canditate['status'] == status['name']:
				canditate['status_id'] = status['id']
		if 'status_id' not in canditate:
			logging.error(f"В статусах нет статуса '{canditate['status']}'")
			return None
	return canditates


def clean_candidate(candidate, cv_data):
	# ухх
	cleaned_candidate = dict()
	with suppress(KeyError, TypeError):
		cleaned_candidate['last_name'] = cv_data['fields']['name']['last']
	with suppress(KeyError, TypeError):
		cleaned_candidate['first_name'] = cv_data['fields']['name']['first']
	with suppress(KeyError, TypeError):
		cleaned_candidate['middle_name'] = cv_data['fields']['name']['middle']
	with suppress(KeyError, TypeError):
		cleaned_candidate['phone'] = cv_data['fields']['phones'][0]
	with suppress(KeyError, TypeError):
		cleaned_candidate['email'] = cv_data['fields']['email']
	with suppress(KeyError, TypeError):
		cleaned_candidate['position'] = cv_data['fields']['position']
	with suppress(KeyError, TypeError):
		cleaned_candidate['company'] = cv_data['fields']['experience'][0]['company']
	with suppress(KeyError, TypeError):
		cleaned_candidate['money'] = cv_data['fields']['salary']
	with suppress(KeyError, TypeError):
		cleaned_candidate['birthday_day'] = cv_data['fields']['birthdate']['day']
	with suppress(KeyError, TypeError):
		cleaned_candidate['birthday_month'] = cv_data['fields']['birthdate']['month']
	with suppress(KeyError, TypeError):
		cleaned_candidate['birthday_year'] = cv_data['fields']['birthdate']['year']
	with suppress(KeyError, TypeError):
		cleaned_candidate['photo'] = cv_data['photo']['id']
	cleaned_candidate['externals'] = \
		[
			{
				"data": {
					"body": cv_data.get("text", "")
				},
				"auth_type": cv_data.get("auth_type", "NATIVE"),
				"files": [
					{
						"id": cv_data.get("id", None)
					}
				],
				"account_source": None
			}
		]
	if 'position' not in cleaned_candidate or cleaned_candidate['position'] is None:
		cleaned_candidate['position'] = candidate['position']
	if 'money' not in cleaned_candidate or cleaned_candidate['money'] is None:
		cleaned_candidate['money'] = candidate['salary']
	with suppress(IndexError):
		if 'last_name' not in cleaned_candidate:
			cleaned_candidate['last_name'] = candidate['name'].split()[0]
		if 'first_name' not in cleaned_candidate:
			cleaned_candidate['first_name'] = candidate['name'].split()[1]
		if 'middle_name' not in cleaned_candidate:
			cleaned_candidate['middle_name'] = candidate['name'].split()[2]
	return cleaned_candidate
